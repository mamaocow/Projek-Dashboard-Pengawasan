#!/usr/bin/env python3
"""
convert.py — Script pembersihan data untuk Dashboard Pengawasan LJK OJK Kalsel 2026.
Mengunduh langsung dari Google Sheets (Published), lalu menghasilkan data.json.

Cara pakai:
  python3 data/convert.py
"""

import openpyxl
import json
import os
import sys
import io
import urllib.request
from datetime import datetime, date, timedelta
from collections import Counter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "data.json")

# ─── Link Google Sheets (Published) ─────────────────────────────────────────
# Ubah ID ini jika link Google Sheets berubah.
# Format: https://docs.google.com/spreadsheets/d/e/XXXX/pubhtml
# → Ganti 'pubhtml' menjadi 'pub?output=xlsx' untuk download Excel langsung.
GSHEET_ID = "2PACX-1vSrQv-NqMkGET_6GNyc16k-JaVe5oyIPNAJyLg6aZkMWNHmpWRdR30uG8tZ_MFqWg"
GSHEET_URL = f"https://docs.google.com/spreadsheets/d/e/{GSHEET_ID}/pub?output=xlsx"
# ─────────────────────────────────────────────────────────────────────────────

BULAN_INDONESIA = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}
BULAN_KE_ANGKA = {v.lower(): k for k, v in BULAN_INDONESIA.items()}


def log(msg): print(f"[INFO] {msg}")
def warn(msg): print(f"[WARN] {msg}", file=sys.stderr)


def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()


def parse_tanggal(value, label=""):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    s = str(value).strip()
    if not s or s.lower() in ('none', 'nan', ''):
        return None

    # "14 Januari 2026" style
    parts = s.split()
    if len(parts) == 3:
        try:
            d, m, y = int(parts[0]), parts[1].lower(), int(parts[2])
            if m in BULAN_KE_ANGKA:
                return date(y, BULAN_KE_ANGKA[m], d)
        except (ValueError, KeyError):
            pass

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:len(fmt)], fmt).date()
        except ValueError:
            continue

    # Excel serial date
    try:
        serial = float(s)
        if 1 < serial < 200000:
            return (date(1899, 12, 30) + timedelta(days=int(serial)))
    except ValueError:
        pass

    warn(f"Gagal parse tanggal '{s}' — {label}")
    return None


def download_workbook():
    log(f"Mengunduh spreadsheet dari Google Sheets...")
    log(f"URL: {GSHEET_URL}")
    req = urllib.request.Request(GSHEET_URL, headers={
        "User-Agent": "Mozilla/5.0"
    })
    with urllib.request.urlopen(req, timeout=30) as r:
        data = r.read()
    log(f"Unduhan selesai ({len(data)//1024} KB).")
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


def main():
    # Download workbook dari Google Sheets
    try:
        wb = download_workbook()
    except Exception as e:
        warn(f"Gagal mengunduh dari Google Sheets: {e}")
        sys.exit(1)

    if "DATA" not in wb.sheetnames:
        warn("Sheet 'DATA' tidak ditemukan!")
        sys.exit(1)

    sheet = wb["DATA"]

    # Row 1 = Header (sesuai inspeksi: spreadsheet sudah di-publish tanpa baris judul)
    header_row = 1
    data_start_row = 2

    headers = [clean_str(c.value) for c in sheet[header_row]]
    log(f"Header ditemukan: {headers[:13]}")

    # Buat peta nama kolom → indeks
    col = {}
    for i, h in enumerate(headers):
        if h:
            col[h] = i

    # Cek kolom wajib
    required = [
        "No", "Nama Bank / Kegiatan Pengawasan", "Nama PUJK", "Kota/Kab",
        "Sektor", "Jenis LJK", "Jenis Kegiatan", "Kuartal (Q)",
        "Tanggal Mulai", "Tanggal Selesai", "Supervisor", "Status Kegiatan"
    ]
    missing = [c for c in required if c not in col]
    if missing:
        warn(f"Kolom tidak ditemukan: {missing}. Kolom yang ada: {list(col.keys())}")

    kegiatan = []
    hari_libur = []
    date_warnings = []

    for row_cells in sheet.iter_rows(min_row=data_start_row, values_only=False):
        row = {name: (row_cells[idx].value if idx < len(row_cells) else None)
               for name, idx in col.items()}

        # Skip baris kosong total
        vals = [v for v in row.values() if v is not None]
        if not vals:
            continue

        no_str = clean_str(row.get("No", "")).replace(".0", "").strip()
        nama_keg = clean_str(row.get("Nama Bank / Kegiatan Pengawasan"))
        nama_pujk = clean_str(row.get("Nama PUJK"))
        kota = clean_str(row.get("Kota/Kab"))
        tgl_mulai_raw = row.get("Tanggal Mulai")

        # Deteksi baris hari libur
        if nama_keg and not nama_pujk and not kota and tgl_mulai_raw is None:
            hari_libur.append({"no": no_str, "keterangan": nama_keg})
            continue

        label = f"No={no_str} '{nama_keg[:40]}'"
        tgl_mulai = parse_tanggal(tgl_mulai_raw, label)
        tgl_selesai = parse_tanggal(row.get("Tanggal Selesai"), label)

        date_warn = False
        if tgl_mulai and tgl_selesai and tgl_selesai < tgl_mulai:
            msg = f"{label}: Tanggal Selesai < Tanggal Mulai"
            warn(msg)
            date_warnings.append(msg)
            date_warn = True

        bulan = BULAN_INDONESIA.get(tgl_mulai.month, "") if tgl_mulai else ""

        # Kuartal: baca dari file atau hitung dari tanggal
        kuartal_raw = clean_str(row.get("Kuartal (Q)"))
        if not kuartal_raw and tgl_mulai:
            q = (tgl_mulai.month - 1) // 3 + 1
            kuartal_raw = f"Q{q}"

        kegiatan.append({
            "no": no_str,
            "namaKegiatan": nama_keg,
            "namaPUJK": nama_pujk,
            "kotaKab": kota,
            "sektor": clean_str(row.get("Sektor")),
            "jenisLJK": clean_str(row.get("Jenis LJK")),
            "jenisKegiatan": clean_str(row.get("Jenis Kegiatan")),
            "kuartal": kuartal_raw,
            "bulan": bulan,
            "tanggalMulai": tgl_mulai.isoformat() if tgl_mulai else None,
            "tanggalSelesai": tgl_selesai.isoformat() if tgl_selesai else None,
            "supervisor": clean_str(row.get("Supervisor")),
            "statusKegiatan": clean_str(row.get("Status Kegiatan")),
            "dateWarning": date_warn,
        })

    # Top 10 PUJK
    counter = Counter(k["namaPUJK"] for k in kegiatan if k["namaPUJK"])
    lookup = {}
    for k in kegiatan:
        if k["namaPUJK"] and k["namaPUJK"] not in lookup:
            lookup[k["namaPUJK"]] = {"sektor": k["sektor"], "jenisLJK": k["jenisLJK"]}

    top10 = [
        {"nama": name, "jumlah": cnt,
         "sektor": lookup.get(name, {}).get("sektor", ""),
         "jenisLJK": lookup.get(name, {}).get("jenisLJK", "")}
        for name, cnt in counter.most_common(10)
    ]

    output = {
        "kegiatan": kegiatan,
        "hariLibur": hari_libur,
        "topPUJK": top10,
        "metadata": {
            "totalKegiatan": len(kegiatan),
            "tanggalGenerate": datetime.now().isoformat(),
            "sumber": GSHEET_URL,
            "warnings": date_warnings,
        },
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    log(f"✅ Berhasil: {len(kegiatan)} kegiatan, {len(hari_libur)} hari libur, {len(top10)} Top PUJK.")
    log(f"Output: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
