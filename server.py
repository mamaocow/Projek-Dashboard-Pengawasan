#!/usr/bin/env python3
"""
server.py — Custom HTTP Server untuk Dashboard Pengawasan LJK OJK Kalsel.
Secara otomatis mengunduh & mengonversi data dari Google Sheets secara real-time
setiap kali browser meminta data/data.json (atau /api/data).
"""

import http.server
import socketserver
import urllib.request
import json
import io
import os
import sys
import openpyxl
from datetime import datetime, date, timedelta
from collections import Counter

PORT = 8000
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_JSON_PATH = os.path.join(SCRIPT_DIR, "data", "data.json")

GSHEET_ID = "2PACX-1vTswl_KO5Flwh1dMg8RJv8wbJnXAmJ13yei15FV8NmiO9O4Rt9RGGcfTFeK9HoynQ"
GSHEET_URL = f"https://docs.google.com/spreadsheets/d/e/{GSHEET_ID}/pub?output=xlsx"

BULAN_INDONESIA = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}
BULAN_KE_ANGKA = {v.lower(): k for k, v in BULAN_INDONESIA.items()}

# English month names (from new spreadsheet)
BULAN_ENGLISH = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def parse_tanggal(value, label=""):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    s = str(value).strip()
    if not s or s.lower() in ('none', 'nan', ''):
        return None

    parts = s.split()
    if len(parts) == 3:
        try:
            d, m, y = int(parts[0]), parts[1].lower(), int(parts[2])
            if m in BULAN_KE_ANGKA:
                return date(y, BULAN_KE_ANGKA[m], d)
        except (ValueError, KeyError):
            pass

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:len(fmt)], fmt).date()
        except ValueError:
            continue

    try:
        serial = float(s)
        if 1 < serial < 200000:
            return (date(1899, 12, 30) + timedelta(days=int(serial)))
    except ValueError:
        pass

    return None

def fetch_and_convert_gsheet():
    """Mengunduh dari Google Sheets dan mengembalikan dict data berformat JSON"""
    req = urllib.request.Request(GSHEET_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=15) as r:
        raw_data = r.read()

    wb = openpyxl.load_workbook(io.BytesIO(raw_data), data_only=True)
    if "DATA" not in wb.sheetnames:
        raise ValueError("Sheet 'DATA' tidak ditemukan")

    sheet = wb["DATA"]
    header_row = 1
    data_start_row = 2

    headers = [clean_str(c.value) for c in sheet[header_row]]
    col = {h: i for i, h in enumerate(headers) if h}

    kegiatan = []
    hari_libur = []
    date_warnings = []

    for row_cells in sheet.iter_rows(min_row=data_start_row, values_only=False):
        row = {name: (row_cells[idx].value if idx < len(row_cells) else None) for name, idx in col.items()}
        if not any(row.values()):
            continue

        no_str = clean_str(row.get("No", "")).replace(".0", "").strip()
        nama_keg = clean_str(row.get("Nama Bank / Kegiatan Pengawasan"))
        nama_pujk = clean_str(row.get("Nama PUJK"))
        kota = clean_str(row.get("Kota/Kab"))
        tgl_mulai_raw = row.get("Tanggal Mulai")

        if nama_keg and not nama_pujk and not kota and tgl_mulai_raw is None:
            hari_libur.append({"no": no_str, "keterangan": nama_keg})
            continue

        label = f"No={no_str} '{nama_keg[:40]}'"
        tgl_mulai = parse_tanggal(tgl_mulai_raw, label)
        tgl_selesai = parse_tanggal(row.get("Tanggal Selesai"), label)

        date_warn = False
        if tgl_mulai and tgl_selesai and tgl_selesai < tgl_mulai:
            date_warnings.append(f"{label}: Tanggal Selesai < Tanggal Mulai")
            date_warn = True

        # Resolve bulan: use Tanggal Mulai month, or parse Bulan column (may be English)
        bulan_raw = clean_str(row.get("Bulan", "")).strip().lower()
        if tgl_mulai:
            bulan = BULAN_INDONESIA.get(tgl_mulai.month, "")
        elif bulan_raw in BULAN_KE_ANGKA:
            bulan = BULAN_INDONESIA.get(BULAN_KE_ANGKA[bulan_raw], "")
        elif bulan_raw in BULAN_ENGLISH:
            bulan = BULAN_INDONESIA.get(BULAN_ENGLISH[bulan_raw], "")
        else:
            bulan = ""

        kuartal_raw = clean_str(row.get("Kuartal (Q)"))
        if not kuartal_raw and tgl_mulai:
            q = (tgl_mulai.month - 1) // 3 + 1
            kuartal_raw = f"Q{q}"

        # Anggota tim
        anggota1 = clean_str(row.get("Anggota 1"))
        anggota2 = clean_str(row.get("Anggota 2"))
        anggota3 = clean_str(row.get("Anggota 3"))
        anggota4 = clean_str(row.get("Anggota 4"))
        anggota = [a for a in [anggota1, anggota2, anggota3, anggota4] if a]

        kegiatan.append({
            "no": no_str,
            "namaKegiatan": nama_keg,
            "namaPUJK": nama_pujk,
            "kotaKab": kota,
            "bidang": clean_str(row.get("Bidang") or row.get("Sektor", "")),
            "jenisLJK": clean_str(row.get("Jenis LJK")),
            "jenisKegiatan": clean_str(row.get("Jenis Kegiatan")),
            "kuartal": kuartal_raw,
            "bulan": bulan,
            "tanggalMulai": tgl_mulai.isoformat() if tgl_mulai else None,
            "tanggalSelesai": tgl_selesai.isoformat() if tgl_selesai else None,
            "supervisor": clean_str(row.get("Supervisor")),
            "statusKegiatan": clean_str(row.get("Status Kegiatan")),
            "anggota": anggota,
            "dateWarning": date_warn,
        })

    counter = Counter(k["namaPUJK"] for k in kegiatan if k["namaPUJK"])
    lookup = {}
    for k in kegiatan:
        if k["namaPUJK"] and k["namaPUJK"] not in lookup:
            lookup[k["namaPUJK"]] = {"bidang": k["bidang"], "jenisLJK": k["jenisLJK"]}

    top10 = [
        {"nama": name, "jumlah": cnt,
         "bidang": lookup.get(name, {}).get("bidang", ""),
         "jenisLJK": lookup.get(name, {}).get("jenisLJK", "")}
        for name, cnt in counter.most_common(10)
    ]

    return {
        "kegiatan": kegiatan,
        "hariLibur": hari_libur,
        "topPUJK": top10,
        "metadata": {
            "totalKegiatan": len(kegiatan),
            "tanggalGenerate": datetime.utcnow().isoformat() + "Z",
            "sumber": GSHEET_URL,
            "warnings": date_warnings,
        },
    }

class CustomHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=SCRIPT_DIR, **kwargs)

    def do_GET(self):
        # Jika browser meminta data.json (termasuk dengan query param ?t=...), ambil data langsung dari Google Sheets
        if self.path.startswith("/data/data.json"):
            try:
                data_dict = fetch_and_convert_gsheet()
                json_bytes = json.dumps(data_dict, ensure_ascii=False, indent=2).encode("utf-8")
                
                # Simpan juga ke file data.json lokal sebagai backup
                try:
                    os.makedirs(os.path.dirname(DATA_JSON_PATH), exist_ok=True)
                    with open(DATA_JSON_PATH, "wb") as f:
                        f.write(json_bytes)
                except Exception:
                    pass

                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(json_bytes)
                print(f"[{datetime.now().strftime('%H:%M:%S')}] ✅ Real-time data fetched from Google Sheets ({len(data_dict['kegiatan'])} kegiatan)")
                return
            except Exception as e:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] ⚠️ Error fetching Google Sheets: {e}, using local fallback")
                # Fallback to local file if available
                if os.path.exists(DATA_JSON_PATH):
                    return super().do_GET()
                else:
                    self.send_error(500, f"Error fetching Google Sheets: {e}")
                    return

        return super().do_GET()

if __name__ == "__main__":
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("", PORT), CustomHandler) as httpd:
        print(f"🚀 Server Dashboard berjalan di http://localhost:{PORT}")
        print(f"🔄 Mode Real-time Google Sheets AKTIF")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nServer dihentikan.")
